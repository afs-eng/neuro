"""Conversor de tabelas XLSM/XLSX do WAIS-III para CSV.

Converte as tabelas normativas do WAIS-III do formato Excel (.xlsm/.xlsx)
para arquivos CSV na estrutura esperada pelo WAIS3NormLoader.

Estrutura de saída:
    raw_to_scaled/
        verbal/  (um CSV por faixa etária)
        execucao/
    composite_scores/  (A3-A9)
    supplementary/     (B1-B7)
    psychometrics/     (5.7, 5.8a, 5.8b, 5.9)
"""
from __future__ import annotations

import csv
import re
import shutil
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent

FAIXAS = ["16-17", "18-19", "20-29", "30-39", "40-49", "50-59", "60-64", "65-89"]

SUBTEST_VERBAL = ["vocabulario", "semelhancas", "aritmetica", "digitos",
                   "informacao", "compreensao", "sequencia_numeros_letras"]
SUBTEST_EXEC = ["completar_figuras", "codigos", "cubos", "raciocinio_matricial",
                 "arranjo_figuras", "procurar_simbolos", "armar_objetos"]
SUBTEST_ALL = SUBTEST_VERBAL + SUBTEST_EXEC

# Mapeamento: nome da aba no XLSM → chave usada no código
A1_SHEETS = {f: f for f in FAIXAS}

# Mapeamento: nome da coluna no XLSM → chave do subteste
COL_MAP_VERBAL = {
    "Vocabulário": "vocabulario",
    "Semelhanças": "semelhancas",
    "Aritmética": "aritmetica",
    "Dígitos": "digitos",
    "Informação": "informacao",
    "Compreensão": "compreensao",
    "Sequência de Números e Letras": "sequencia_numeros_letras",
}
COL_MAP_EXEC = {
    "Completar Figuras": "completar_figuras",
    "Códigos": "codigos",
    "Cubos": "cubos",
    "Raciocínio Matricial": "raciocinio_matricial",
    "Arranjo de Figuras": "arranjo_figuras",
    "Procurar Símbolos": "procurar_simbolos",
    "Armar Objetos": "armar_objetos",
}


def parse_range(val: str) -> list[tuple[int, int]]:
    """Parse '0–2' → [(0,2)], '—' → [], '1' → [(1,1)]."""
    val = str(val).strip()
    if val in ("—", "", "nan", "None"):
        return []
    parts = val.replace("–", "-").split("-")
    try:
        if len(parts) == 1:
            v = int(parts[0])
            return [(v, v)]
        elif len(parts) == 2:
            return [(int(parts[0]), int(parts[1]))]
    except ValueError:
        pass
    return []


def expand_range_to_raw_scores(val: str) -> list[int]:
    """Expande um intervalo de raw scores em lista de ints."""
    ranges = parse_range(val)
    result = []
    for lo, hi in ranges:
        result.extend(range(lo, hi + 1))
    return result


def convert_a1_to_csv():
    """Converte Tabela A.1 em CSVs raw_to_scaled/verbal/ e execucao/."""
    wb = openpyxl.load_workbook(BASE / "tabela_a1.xlsm", data_only=True)

    for faixa, sheet_name in A1_SHEETS.items():
        ws = wb[sheet_name]

        # Coletar todas as linhas não-vazias
        all_rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]

        # Encontrar onde começam as seções "Subtestes Verbais" e "Subtestes de Execução"
        verbal_section = False
        exec_section = False
        data_rows = []

        for row in all_rows:
            first = str(row[0]).strip() if row[0] else ""
            if "Subtestes Verbais" in first:
                verbal_section = True
                exec_section = False
                continue
            if "Subtestes de Execução" in first:
                verbal_section = False
                exec_section = True
                continue
            if verbal_section or exec_section:
                # Linha de dados: primeiro elemento é escore ponderado
                try:
                    scaled = int(first)
                    data_rows.append((scaled, row, verbal_section))
                except (ValueError, TypeError):
                    # Linha de cabeçalho de subtestes
                    continue

        # Processar verbal
        verbal_out_dir = BASE / "raw_to_scaled" / "verbal"
        verbal_out_dir.mkdir(parents=True, exist_ok=True)
        _write_a1_domain(data_rows, verbal_section=True, col_map=COL_MAP_VERBAL,
                         subtests=SUBTEST_VERBAL, out_dir=verbal_out_dir, faixa=faixa)

        # Processar execucao
        exec_out_dir = BASE / "raw_to_scaled" / "execucao"
        exec_out_dir.mkdir(parents=True, exist_ok=True)
        _write_a1_domain(data_rows, verbal_section=False, col_map=COL_MAP_EXEC,
                         subtests=SUBTEST_EXEC, out_dir=exec_out_dir, faixa=faixa)

    print("✓ A1 convertido: raw_to_scaled/[verbal|execucao]/")


def _write_a1_domain(data_rows, verbal_section, col_map, subtests, out_dir, faixa):
    """Escreve o CSV para um domínio."""
    # Primeiro, expandir cada linha de dados em múltiplas linhas de raw score
    expanded_rows: list[dict] = []
    for scaled, row, is_verbal in data_rows:
        if is_verbal != verbal_section:
            continue
        # Mapear valores da linha para os subtestes corretos
        # A sheet tem colunas: EscPond, Vocab, Semelh, Arit, Dig, Info, Comp, SeqNL (verbal)
        # ou: EscPond, CompFig, Cod, Cubos, RacMat, ArrFig, ProcSim, ArmObj (exec)
        values = list(row)
        # values[0] = escore ponderado
        # values[1..] = valores para cada subteste
        for idx, subtest_key in enumerate(subtests):
            cell_val = values[idx + 1] if idx + 1 < len(values) else None
            if cell_val is None:
                continue
            raw_scores = expand_range_to_raw_scores(str(cell_val))
            for raw in raw_scores:
                expanded_rows.append({
                    "raw_score": raw,
                    subtest_key: scaled,
                })

    if not expanded_rows:
        return

    # Agora pivotar: agrupar por raw_score e横向
    by_raw: dict[int, dict] = {}
    for r in expanded_rows:
        raw = r["raw_score"]
        if raw not in by_raw:
            by_raw[raw] = {"raw_score": raw}
        for k, v in r.items():
            if k != "raw_score":
                by_raw[raw][k] = v

    # Escrever CSV
    out_path = out_dir / f"idade_{faixa}.csv"
    fieldnames = ["raw_score"] + subtests
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for raw in sorted(by_raw.keys()):
            writer.writerow(by_raw[raw])

    print(f"  ✓ {out_path.name}")


def _read_xlsm_sheet(path: Path, sheet_idx: int = 0) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True)
    name = wb.sheetnames[sheet_idx]
    ws = wb[name]
    return [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]


def convert_composite_scores():
    """Converte Tabelas A3-A9 (QI Verbal, QI Execução, QI Total, ICV, IOP, IMO, IVP)."""
    out_dir = BASE / "composite_scores"
    out_dir.mkdir(parents=True, exist_ok=True)

    mappings = [
        ("tabela_a3.xlsm", "qi_verbal.csv", None),
        ("tabela_a4.xlsm", "qi_execucao.csv", None),
        ("tabela_a5.xlsm", "qi_total.csv", None),
        ("tabela_a6.xlsm", "compreensao_verbal.csv", None),
        ("tabela_a7.xlsm", "organizacao_perceptual.csv", None),
        ("tabela_a8.xlsm", "memoria_operacional.csv", None),
        ("tabela_a9.xlsm", "velocidade_processamento.csv", None),
    ]

    for xlsm, csv_name, sheet_override in mappings:
        rows = _read_xlsm_sheet(BASE / xlsm)
        out_path = out_dir / csv_name
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["soma_ponderada", "pontuacao_composta",
                                                    "percentil", "ic_90", "ic_95", "classificacao"])
            writer.writeheader()
            for row in rows:
                if len(row) < 5:
                    continue
                try:
                    soma = int(row[0])
                    qi = int(row[1])
                    percentile = str(row[2]) if row[2] is not None else ""
                    ic90 = str(row[3]) if row[3] is not None else ""
                    ic95 = str(row[4]) if row[4] is not None else ""
                    writer.writerow({
                        "soma_ponderada": soma,
                        "pontuacao_composta": qi,
                        "percentil": percentile,
                        "ic_90": ic90,
                        "ic_95": ic95,
                        "classificacao": "",
                    })
                except (ValueError, TypeError):
                    continue
        print(f"  ✓ {csv_name}")

    print("✓ A3-A9 convertidos: composite_scores/")


def convert_supplementary():
    """Converte Tabelas B1-B7."""
    out_dir = BASE / "supplementary"
    out_dir.mkdir(parents=True, exist_ok=True)

    b_mappings = [
        ("tabela_b1.xlsm", "b1_diferencas_qi_indices_significancia.csv", 0),
        ("tabela_b2.xlsm", "b2_frequencia_diferencas_qi_indices.csv", 0),
        ("tabela_b3.xlsm", "b3_diferencas_subteste_media.csv", 0),
        ("tabela_b4.xlsm", "b4_diferencas_entre_subtestes.csv", 0),
        ("tabela_b5.xlsm", "b5_dispersao_subtestes.csv", 0),
        ("tabela_b6.xlsm", "b6_digitos_ordem_direta_inversa.csv", 0),
        ("tabela_b7.xlsm", "b7_diferenca_digitos_direta_inversa.csv", 0),
    ]

    for xlsm, csv_name, sheet_idx in b_mappings:
        wb = openpyxl.load_workbook(BASE / xlsm, data_only=True)
        # Para B3, pode ter múltiplas sheets
        sheets_to_process = wb.sheetnames if xlsm == "tabela_b3.xlsm" else [wb.sheetnames[sheet_idx]]

        all_rows = []
        for sn in sheets_to_process:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    all_rows.append(row)

        out_path = out_dir / csv_name
        max_cols = max((len(r) for r in all_rows), default=0)
        fieldnames = [f"col_{i}" for i in range(max_cols)]
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in all_rows:
                padded = list(row) + [""] * (max_cols - len(row))
                writer.writerow({f"col_{i}": v for i, v in enumerate(padded)})
        print(f"  ✓ {csv_name}")

    print("✓ B1-B7 convertidos: supplementary/")


def convert_psychometrics():
    """Converte Tabelas 5.7, 5.8a, 5.8b, 5.9."""
    out_dir = BASE / "psychometrics"
    out_dir.mkdir(parents=True, exist_ok=True)

    mappings = [
        ("tabela_5_7.xlsm", "consistencia_interna.csv"),
        ("tabela_5_8a.xlsm", "estabilidade_teste_reteste.csv"),
        ("tabela_5_8b.xlsm", "amostra_reteste.csv"),
        ("tabela_5_9.xlsm", "erro_padrao_medida.csv"),
    ]

    for xlsm, csv_name in mappings:
        rows = _read_xlsm_sheet(BASE / xlsm)
        out_path = out_dir / csv_name
        max_cols = max((len(r) for r in rows), default=0)
        fieldnames = [f"col_{i}" for i in range(max_cols)]
        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                padded = list(row) + [""] * (max_cols - len(row))
                writer.writerow({f"col_{i}": v for i, v in enumerate(padded)})
        print(f"  ✓ {csv_name}")

    print("✓ Tabelas psicométricas (5.7-5.9) convertidas: psychometrics/")


def main():
    print("=== Convertendo tabelas WAIS-III XLSM → CSV ===\n")
    convert_a1_to_csv()
    print()
    convert_composite_scores()
    print()
    convert_supplementary()
    print()
    convert_psychometrics()
    print("\n=== Conversão concluída ===")


if __name__ == "__main__":
    main()
